"""Выгрузка базы в Excel (.xlsx) для администратора."""
from __future__ import annotations

import io
from datetime import datetime

from sqlalchemy import select

from config import ADMIN_IDS
from db.database import AsyncSessionLocal
from db.models import CatalogItem, Category, Order, User

ORDERS_HEADERS = [
    "№ заказа", "Дата создания", "Статус", "Клиент (ФИО)", "Телефон", "Город",
    "Пункт 5Post", "Комментарий", "Состав заказа", "Растения, ₽", "Доставка, ₽",
    "Итого, ₽", "Предоплата, ₽", "Остаток, ₽", "Трек-номер",
    "Telegram ID", "Username",
]

CLIENTS_HEADERS = ["Telegram ID", "Username", "ФИО", "Телефон", "Город", "Пункт 5Post", "Дата регистрации", "Заказов всего"]

ITEMS_HEADERS = ["Категория", "№ фото", "Название", "Тип", "Цена, ₽", "Остаток", "Активна"]


def _order_items_str(order: Order) -> str:
    return "; ".join(
        f"{oi.catalog_item.category.name} №{oi.catalog_item.photo_number} × {oi.quantity} шт. по {oi.price_at_order:.0f} ₽"
        for oi in order.items
    )


async def build_export_xlsx() -> bytes:
    from openpyxl import Workbook
    from openpyxl.styles import Font

    wb = Workbook()
    bold = Font(bold=True)

    async with AsyncSessionLocal() as s:
        orders = list((await s.execute(select(Order).order_by(Order.created_at.desc()))).scalars().all())
        users = list((await s.execute(select(User).order_by(User.created_at.desc()))).scalars().all())
        items = list((await s.execute(select(CatalogItem).order_by(CatalogItem.category_id, CatalogItem.photo_number))).scalars().all())

    ws = wb.active
    ws.title = "Заказы"
    ws.append(ORDERS_HEADERS)
    for cell in ws[1]:
        cell.font = bold
    for o in orders:
        ws.append([
            o.id, o.created_at.strftime("%Y-%m-%d %H:%M"), o.status.value,
            o.full_name or "", o.phone or "", o.city or "", o.pickup_point or "",
            o.comment or "", _order_items_str(o),
            o.plants_cost, o.delivery_cost, o.total_cost, o.prepayment, o.remainder,
            o.track_number or "", o.user.telegram_id, o.user.username or "",
        ])
    ws.column_dimensions["I"].width = 60

    ws2 = wb.create_sheet("Клиенты")
    ws2.append(CLIENTS_HEADERS)
    for cell in ws2[1]:
        cell.font = bold
    for u in users:
        ws2.append([
            u.telegram_id, u.username or "", u.full_name or "", u.phone or "",
            u.city or "", u.pickup_point or "",
            u.created_at.strftime("%Y-%m-%d") if u.created_at else "",
            len(u.orders),
        ])

    ws3 = wb.create_sheet("Позиции каталога")
    ws3.append(ITEMS_HEADERS)
    for cell in ws3[1]:
        cell.font = bold
    for i in items:
        ws3.append([
            i.category.name, i.photo_number, i.title or "",
            "саженец" if i.kind == "sapling" else "товар",
            i.price, i.stock, "да" if i.is_active else "нет",
        ])

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


EMPLOYEE_CLIENTS_HEADERS = [
    "ФИО", "Телефон", "Telegram ID", "Username", "Источник", "Дата регистрации",
    "Заказов", "Сумма, ₽",
]

EMPLOYEE_ORDERS_HEADERS = [
    "№ заказа", "Дата", "Статус", "Клиент", "Телефон", "Город", "Пункт 5Post",
    "Сумма, ₽", "Предоплата, ₽", "Остаток, ₽", "Источник", "Трек-номер",
]


async def build_employee_export_xlsx(employee_id: int) -> bytes | None:
    """Выгрузка клиентов и заказов конкретного сотрудника."""
    from openpyxl import Workbook
    from openpyxl.styles import Font

    from db.crud import get_employee_stats

    stats = await get_employee_stats(employee_id)
    if not stats["clients"] and not stats["orders"]:
        return None

    wb = Workbook()
    bold = Font(bold=True)

    orders_by_user: dict[int, list] = {}
    for order in stats["orders"]:
        orders_by_user.setdefault(order.user_id, []).append(order)

    ws = wb.active
    ws.title = "Клиенты"
    ws.append(EMPLOYEE_CLIENTS_HEADERS)
    for cell in ws[1]:
        cell.font = bold
    for user in stats["clients"]:
        user_orders = orders_by_user.get(user.id, [])
        total = sum(o.total_cost for o in user_orders)
        ws.append([
            user.full_name or user.username or "",
            user.phone or "",
            user.telegram_id or "",
            user.username or "",
            user.source or "",
            user.created_at.strftime("%Y-%m-%d") if user.created_at else "",
            len(user_orders),
            total,
        ])

    ws2 = wb.create_sheet("Заказы")
    ws2.append(EMPLOYEE_ORDERS_HEADERS)
    for cell in ws2[1]:
        cell.font = bold
    for order in stats["orders"]:
        ws2.append([
            order.id,
            order.created_at.strftime("%Y-%m-%d %H:%M") if order.created_at else "",
            order.status.value,
            order.full_name or "",
            order.phone or "",
            order.city or "",
            order.pickup_point or "",
            order.total_cost,
            order.prepayment,
            order.remainder,
            order.source or "",
            order.track_number or "",
        ])

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


async def send_export_to_admin(bot, caption: str | None = None) -> None:
    """Собирает выгрузку и отправляет файл всем админам."""
    from aiogram.types import BufferedInputFile

    data = await build_export_xlsx()
    filename = f"pitomnik_{datetime.now().strftime('%Y%m%d_%H%M')}.xlsx"
    for admin_id in ADMIN_IDS:
        try:
            await bot.send_document(admin_id, BufferedInputFile(data, filename=filename), caption=caption or "📊 Выгрузка базы")
        except Exception as exc:
            print(f"[export] Не удалось отправить выгрузку админу {admin_id}: {exc}")
